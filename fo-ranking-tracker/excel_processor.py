from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from price_fetcher import PriceFetcher, get_fridays

# Source columns in uploaded file
SRC_SCRIPT = 1   # A
SRC_SECTOR = 3   # C
SRC_SEGMENT = 4  # D

# Output columns in clean workbook
OUT_SCRIPT = 1
OUT_SECTOR = 2
OUT_SEGMENT = 3
OUT_BASE = 4
OUT_CURRENT = 5
OUT_POINTS = 6
OUT_PCT = 7
OUT_RANK = 8
OUT_FRIDAY_START = 9  # I

TOP_N_FRIDAY_RANK = 20
FRIDAY_SUMMARY_SHEET = "Friday Top 20"

SCRIPT_HEADERS = ("script", "symbol", "ticker")

# Rank-based colors (matches your Excel: 1-20 blue, 21-40 green, 41-60 yellow, 61-90 brown)
RANK_COLOR_BANDS: Tuple[Tuple[int, int, str], ...] = (
    (1, 20, "BDD7EE"),   # light blue
    (21, 40, "C6EFCE"),  # light green
    (41, 60, "FFEB9C"),  # yellow
    (61, 90, "F4B084"),  # light brown
)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _read_headers(ws, header_row: int = 1) -> Dict[int, str]:
    headers: Dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        headers[col] = _normalize_header(ws.cell(header_row, col).value)
    return headers


def _find_script_column(headers: Dict[int, str]) -> int:
    for col, header in headers.items():
        if header in SCRIPT_HEADERS:
            return col
    return SRC_SCRIPT


def _detect_data_rows(ws, script_col: int, start_row: int = 2) -> Tuple[int, int]:
    first_row = start_row
    last_row = start_row - 1

    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, script_col).value
        if value is None or str(value).strip() == "":
            continue
        if last_row < start_row:
            first_row = row
        last_row = row

    if last_row < start_row:
        raise ValueError("No stock symbols found in column A (Script).")

    return first_row, last_row


def _rank_desc(values: List[Optional[float]]) -> List[Optional[int]]:
    indexed = [(idx, val) for idx, val in enumerate(values) if val is not None]
    indexed.sort(key=lambda item: item[1], reverse=True)

    ranks: List[Optional[int]] = [None] * len(values)
    rank = 1
    for idx, _ in indexed:
        ranks[idx] = rank
        rank += 1
    return ranks


def _top_n_symbols(
    symbols: List[str],
    values: List[Optional[float]],
    n: int = TOP_N_FRIDAY_RANK,
) -> List[str]:
    ranked = [
        (symbols[idx], value)
        for idx, value in enumerate(values)
        if value is not None and str(symbols[idx]).strip()
    ]
    ranked.sort(key=lambda item: item[1], reverse=True)
    return [symbol for symbol, _ in ranked[:n]]


def _pct_change(base: Optional[float], current: Optional[float]) -> Optional[float]:
    if base is None or current is None or base == 0:
        return None
    return ((current - base) / base) * 100.0


def _fill_for_rank(rank: Optional[object]) -> Optional[PatternFill]:
    if rank is None:
        return None
    try:
        value = int(rank)
    except (TypeError, ValueError):
        return None

    for low, high, color in RANK_COLOR_BANDS:
        if low <= value <= high:
            return PatternFill(fill_type="solid", fgColor=color)
    return None


def _apply_rank_fill(cell, rank: Optional[object]) -> None:
    fill = _fill_for_rank(rank)
    if fill is not None:
        cell.fill = fill


def _build_friday_summary_sheet(
    wb: Workbook,
    fridays: List[date],
    symbols: List[str],
    base_prices: List[Optional[float]],
    fetcher: PriceFetcher,
) -> None:
    if FRIDAY_SUMMARY_SHEET in wb.sheetnames:
        del wb[FRIDAY_SUMMARY_SHEET]

    summary = wb.create_sheet(FRIDAY_SUMMARY_SHEET)

    for col_idx, friday in enumerate(fridays, start=1):
        summary.cell(1, col_idx, friday.strftime("%d-%m-%Y"))

        friday_pcts: List[Optional[float]] = []
        for base_price, symbol in zip(base_prices, symbols):
            friday_price = fetcher.price_on(symbol, friday)
            friday_pcts.append(_pct_change(base_price, friday_price))

        top_symbols = _top_n_symbols(symbols, friday_pcts, TOP_N_FRIDAY_RANK)
        for row_offset, symbol in enumerate(top_symbols, start=2):
            cell = summary.cell(row_offset, col_idx, symbol)
            _apply_rank_fill(cell, row_offset - 1)


def process_workbook(
    file_bytes: bytes,
    base_date: date,
    today: Optional[date] = None,
) -> bytes:
    today = today or date.today()
    if base_date > today:
        raise ValueError("Base date cannot be after today.")

    src_values = load_workbook(BytesIO(file_bytes), data_only=True).active

    headers = _read_headers(src_values)
    script_col = _find_script_column(headers)
    data_start_row, data_end_row = _detect_data_rows(src_values, script_col)

    symbols: List[str] = []
    sectors: List[object] = []
    segments: List[object] = []

    for row in range(data_start_row, data_end_row + 1):
        symbol = src_values.cell(row, script_col).value
        if symbol is None or str(symbol).strip() == "":
            continue
        symbols.append(str(symbol).strip())
        sectors.append(src_values.cell(row, SRC_SECTOR).value)
        segments.append(src_values.cell(row, SRC_SEGMENT).value)

    fridays = get_fridays(base_date, today)
    fetcher = PriceFetcher(symbols, base_date, today)
    fetcher.load()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    base_label = f"Base Price ({base_date.strftime('%d-%b-%Y')})"
    ws.cell(1, OUT_SCRIPT, "Script")
    ws.cell(1, OUT_SECTOR, "Sector")
    ws.cell(1, OUT_SEGMENT, "Segment")
    ws.cell(1, OUT_BASE, base_label)
    ws.cell(1, OUT_CURRENT, "Current Price")
    ws.cell(1, OUT_POINTS, "Points")
    ws.cell(1, OUT_PCT, "% Diff")
    ws.cell(1, OUT_RANK, "Rank")

    friday_col_map: Dict[date, int] = {}
    for idx, friday in enumerate(fridays):
        col = OUT_FRIDAY_START + idx
        ws.cell(1, col, friday.strftime("%d-%b-%Y"))
        friday_col_map[friday] = col

    base_prices: List[Optional[float]] = []
    pct_diffs: List[Optional[float]] = []

    for out_row, (symbol, sector, segment) in enumerate(
        zip(symbols, sectors, segments),
        start=2,
    ):
        base_price = fetcher.price_on(symbol, base_date)
        current_price = fetcher.price_on(symbol, today)

        base_prices.append(base_price)

        points = None
        if base_price is not None and current_price is not None:
            points = current_price - base_price

        pct_diff = _pct_change(base_price, current_price)
        pct_diffs.append(pct_diff)

        ws.cell(out_row, OUT_SCRIPT, symbol)
        ws.cell(out_row, OUT_SECTOR, sector)
        ws.cell(out_row, OUT_SEGMENT, segment)
        ws.cell(out_row, OUT_BASE, base_price)
        ws.cell(out_row, OUT_CURRENT, current_price)
        ws.cell(out_row, OUT_POINTS, points)
        ws.cell(out_row, OUT_PCT, pct_diff)

    today_ranks = _rank_desc(pct_diffs)
    for out_row, rank in enumerate(today_ranks, start=2):
        rank_cell = ws.cell(out_row, OUT_RANK, rank)
        _apply_rank_fill(rank_cell, rank)

    for friday in fridays:
        friday_pcts: List[Optional[float]] = []
        for base_price, symbol in zip(base_prices, symbols):
            friday_price = fetcher.price_on(symbol, friday)
            friday_pcts.append(_pct_change(base_price, friday_price))

        friday_ranks = _rank_desc(friday_pcts)
        col = friday_col_map[friday]

        for out_row, rank in enumerate(friday_ranks, start=2):
            friday_cell = ws.cell(out_row, col, rank)
            _apply_rank_fill(friday_cell, rank)

    _build_friday_summary_sheet(wb, fridays, symbols, base_prices, fetcher)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
