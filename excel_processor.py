from __future__ import annotations

import re
from copy import copy
from datetime import date
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from price_fetcher import PriceFetcher, get_fridays, trading_date_for

# Keep uploaded columns A, B, C as-is. Calculations start at D.
KEEP_COLS = (1, 2, 3)
SRC_SCRIPT = 1

OUT_BASE = 4
OUT_CURRENT = 5
OUT_POINTS = 6
OUT_PCT = 7
OUT_RANK = 8
OUT_FRIDAY_START = 9

TOP_N_FRIDAY_RANK = 20
FRIDAY_SUMMARY_SHEET = "Friday Top 20"
TOP_GRID_START_ROW = 2
TOP_GRID_END_ROW = 21

FILL_ENTRY = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_EXIT = PatternFill(fill_type="solid", fgColor="FFC7CE")

SCRIPT_HEADERS = ("script", "symbol", "ticker")

RANK_COLOR_BANDS: Tuple[Tuple[int, int, str], ...] = (
    (1, 20, "BDD7EE"),
    (21, 40, "C6EFCE"),
    (41, 60, "FFEB9C"),
    (61, 90, "F4B084"),
)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _read_headers(ws, header_row: int = 1) -> Dict[int, str]:
    headers: Dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        headers[col] = _normalize_header(ws.cell(header_row, col).value)
    return headers


def _find_script_column(headers: Dict[int, str]) -> int:
    for col, header in headers.items():
        if header in SCRIPT_HEADERS:
            return col
    return SRC_SCRIPT


def _detect_data_rows(ws, script_col: int, start_row: int = 2) -> Tuple[int, int]:
    first_row = start_row
    last_row = start_row - 1

    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, script_col).value
        if value is None or str(value).strip() == "":
            continue
        if last_row < start_row:
            first_row = row
        last_row = row

    if last_row < start_row:
        raise ValueError("No stock symbols found in column A (Script).")

    return first_row, last_row


def _copy_cell(source, target) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def _copy_header_labels(src_ws, out_ws) -> None:
    for col in KEEP_COLS:
        label = src_ws.cell(1, col).value
        if label is not None and str(label).strip():
            out_ws.cell(1, col, label)


def _rank_desc(values: List[Optional[float]]) -> List[Optional[int]]:
    indexed = [(idx, val) for idx, val in enumerate(values) if val is not None]
    indexed.sort(key=lambda item: item[1], reverse=True)

    ranks: List[Optional[int]] = [None] * len(values)
    rank = 1
    for idx, _ in indexed:
        ranks[idx] = rank
        rank += 1
    return ranks


def _top_n_symbols(
    symbols: List[str],
    values: List[Optional[float]],
    n: int = TOP_N_FRIDAY_RANK,
) -> List[str]:
    ranked = [
        (symbols[idx], value)
        for idx, value in enumerate(values)
        if value is not None and str(symbols[idx]).strip()
    ]
    ranked.sort(key=lambda item: item[1], reverse=True)
    return [symbol for symbol, _ in ranked[:n]]


def _friday_header_label(friday: date) -> str:
    trading = trading_date_for(friday)
    if trading != friday:
        return f"{friday.strftime('%d-%b-%Y')} ({trading.strftime('%d-%b-%Y')})"
    return friday.strftime("%d-%b-%Y")


def _friday_summary_label(friday: date) -> str:
    trading = trading_date_for(friday)
    if trading != friday:
        return f"{friday.strftime('%d-%m-%Y')} ({trading.strftime('%d-%m-%Y')})"
    return friday.strftime("%d-%m-%Y")


def _pct_diff(base: Optional[float], price: Optional[float]) -> Optional[float]:
    """% Diff = (points / base price) * 100, where points = price - base."""
    if base is None or price is None or base == 0:
        return None
    points = round(price - base, 2)
    return (points / base) * 100.0


def _fill_for_rank(rank: Optional[object]) -> Optional[PatternFill]:
    if rank is None:
        return None
    try:
        value = int(rank)
    except (TypeError, ValueError):
        return None

    for low, high, color in RANK_COLOR_BANDS:
        if low <= value <= high:
            return PatternFill(fill_type="solid", fgColor=color)
    return None


def _apply_rank_fill(cell, rank: Optional[object]) -> None:
    fill = _fill_for_rank(rank)
    if fill is not None:
        cell.fill = fill


def _build_friday_summary_sheet(
    wb: Workbook,
    fridays: List[date],
    symbols: List[str],
    base_prices: List[Optional[float]],
    fetcher: PriceFetcher,
) -> None:
    if FRIDAY_SUMMARY_SHEET in wb.sheetnames:
        del wb[FRIDAY_SUMMARY_SHEET]

    summary = wb.create_sheet(FRIDAY_SUMMARY_SHEET)
    friday_tops: List[List[str]] = []

    for col_idx, friday in enumerate(fridays, start=1):
        summary.cell(1, col_idx, _friday_summary_label(friday))

        friday_pcts: List[Optional[float]] = []
        for base_price, symbol in zip(base_prices, symbols):
            friday_price = fetcher.price_on(symbol, friday)
            friday_pcts.append(_pct_diff(base_price, friday_price))

        top_symbols = _top_n_symbols(symbols, friday_pcts, TOP_N_FRIDAY_RANK)
        friday_tops.append(top_symbols)

        for row_offset, symbol in enumerate(top_symbols, start=TOP_GRID_START_ROW):
            summary.cell(row_offset, col_idx, symbol)

    exits_per_col: List[List[str]] = []
    entries_per_col: List[List[str]] = []
    max_exits = 0
    max_entries = 0

    for col_idx, top_symbols in enumerate(friday_tops):
        prev_set = set(friday_tops[col_idx - 1]) if col_idx > 0 else set()
        curr_set = set(top_symbols)

        entries = [symbol for symbol in top_symbols if symbol not in prev_set]
        exits = sorted(prev_set - curr_set)

        entries_per_col.append(entries)
        exits_per_col.append(exits)
        max_exits = max(max_exits, len(exits))
        max_entries = max(max_entries, len(entries))

        for row_offset, symbol in enumerate(top_symbols, start=TOP_GRID_START_ROW):
            cell = summary.cell(row_offset, col_idx + 1)
            if col_idx == 0 or symbol in entries:
                cell.fill = FILL_ENTRY

    # If column B (or any week) has exits, mark those stocks red in the
    # previous column's top-20 grid (e.g. exits in B -> red in column A).
    for col_idx, exits in enumerate(exits_per_col):
        if col_idx == 0 or not exits:
            continue
        prev_col = col_idx
        prev_top = friday_tops[col_idx - 1]
        exit_set = set(exits)
        for row_offset, symbol in enumerate(prev_top, start=TOP_GRID_START_ROW):
            if symbol in exit_set:
                summary.cell(row_offset, prev_col).fill = FILL_EXIT

    exit_header_row = TOP_GRID_END_ROW + 2
    exit_start_row = exit_header_row + 1
    exit_block_rows = max(max_exits, 1)

    for col_idx, exits in enumerate(exits_per_col, start=1):
        summary.cell(exit_header_row, col_idx, "Exit")
        for offset, symbol in enumerate(exits):
            summary.cell(exit_start_row + offset, col_idx, symbol)

    entry_header_row = exit_start_row + exit_block_rows + 1
    entry_start_row = entry_header_row + 1
    entry_block_rows = max(max_entries, 1)

    for col_idx, entries in enumerate(entries_per_col, start=1):
        summary.cell(entry_header_row, col_idx, "Entry")
        for offset, symbol in enumerate(entries):
            cell = summary.cell(entry_start_row + offset, col_idx, symbol)
            cell.fill = FILL_ENTRY


def process_workbook(
    file_bytes: bytes,
    base_date: date,
    today: Optional[date] = None,
) -> bytes:
    today = today or date.today()
    if base_date > today:
        raise ValueError("Base date cannot be after today.")

    src_wb = load_workbook(BytesIO(file_bytes), data_only=False)
    src_ws = src_wb.active
    src_values = load_workbook(BytesIO(file_bytes), data_only=True).active

    headers = _read_headers(src_values)
    script_col = _find_script_column(headers)
    data_start_row, data_end_row = _detect_data_rows(src_values, script_col)

    symbols: List[str] = []
    source_rows: List[int] = []

    for row in range(data_start_row, data_end_row + 1):
        symbol = src_values.cell(row, script_col).value
        if symbol is None or str(symbol).strip() == "":
            continue
        symbols.append(str(symbol).strip())
        source_rows.append(row)

    fridays = get_fridays(base_date, today)
    fetcher = PriceFetcher(symbols, base_date, today)
    fetcher.load()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    _copy_header_labels(src_ws, ws)

    base_label = f"Base Price ({base_date.strftime('%d-%b-%Y')})"
    ws.cell(1, OUT_BASE, base_label)
    ws.cell(1, OUT_CURRENT, "Current Price")
    ws.cell(1, OUT_POINTS, "Points")
    ws.cell(1, OUT_PCT, "% Diff")
    ws.cell(1, OUT_RANK, "Rank")

    friday_col_map: Dict[date, int] = {}
    for idx, friday in enumerate(fridays):
        col = OUT_FRIDAY_START + idx
        ws.cell(1, col, _friday_header_label(friday))
        friday_col_map[friday] = col

    base_prices: List[Optional[float]] = []
    pct_diffs: List[Optional[float]] = []

    for out_row, (symbol, src_row) in enumerate(zip(symbols, source_rows), start=2):
        for col in KEEP_COLS:
            _copy_cell(src_ws.cell(src_row, col), ws.cell(out_row, col))

        base_price = fetcher.price_on(symbol, base_date)
        current_price = fetcher.price_on(symbol, today)

        base_prices.append(base_price)

        points = None
        if base_price is not None and current_price is not None:
            points = round(current_price - base_price, 2)

        pct_diff = _pct_diff(base_price, current_price)
        pct_diffs.append(pct_diff)

        base_cell = ws.cell(out_row, OUT_BASE, base_price)
        base_cell.number_format = "0.00"
        current_cell = ws.cell(out_row, OUT_CURRENT, current_price)
        current_cell.number_format = "0.00"
        points_cell = ws.cell(out_row, OUT_POINTS, points)
        points_cell.number_format = "0.00"
        pct_cell = ws.cell(out_row, OUT_PCT, pct_diff)
        pct_cell.number_format = "0.00"

    today_ranks = _rank_desc(pct_diffs)
    for out_row, rank in enumerate(today_ranks, start=2):
        rank_cell = ws.cell(out_row, OUT_RANK, rank)
        _apply_rank_fill(rank_cell, rank)

    for friday in fridays:
        friday_pcts: List[Optional[float]] = []
        for base_price, symbol in zip(base_prices, symbols):
            friday_price = fetcher.price_on(symbol, friday)
            friday_pcts.append(_pct_diff(base_price, friday_price))

        friday_ranks = _rank_desc(friday_pcts)
        col = friday_col_map[friday]

        for out_row, rank in enumerate(friday_ranks, start=2):
            friday_cell = ws.cell(out_row, col, rank)
            _apply_rank_fill(friday_cell, rank)

    _build_friday_summary_sheet(wb, fridays, symbols, base_prices, fetcher)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
